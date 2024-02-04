# 必要なライブラリをインポート
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from getpass import getpass
import time
import os
import pandas as pd
import glob
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook

# エクセルを立ち上げる
filepath = path = "/Users/kouheitakahashi/excel_auto/在庫管理表/在庫リスト.xlsx"

# Excelファイルを開く
wb = load_workbook(filename=filepath, data_only=True)

# 最初のシートを選択
ws = wb.active

# B列の各セルをチェック
print(ws.max_row)
for row in range(4, ws.max_row + 1):
    print("welcome2")
    cell = ws['B{}'.format(row)]
    print(cell.value)
    # フラグが立っている場合は申請を行う
    if cell.value == 1:
        # ここで申請を行うコードを実行
        print("Apply for row {}".format(row))

    else:
        print("test")
        # フラグが立っていない場合は繰り返しを終了
