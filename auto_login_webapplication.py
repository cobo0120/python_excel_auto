# 必要なライブラリをインポート
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from getpass import getpass
import time
import os
import pandas as pd
import glob
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
from datetime import datetime, timedelta
# 別のプログラム上でタイマー設定で起動（exeファイルを設定し起動）朝の9時に読み進める
# エクセルを見にいく（python）どこのエクセルを見にいくのか難しいため、あらかじめディレクトリ指定するか。exeファイルを作成するか。

# フォルダに保存されている指定のエクセルファイルを参照する


# def list_files_recursively(directory):
#     return glob.glob(os.path.join(directory, '**/*.xlsx'), recursive=True)


# path = '/Users/kouheitakahashi/excel_auto/在庫管理表/在庫リスト.xlsx'
# excel_files = list_files_recursively(path)

# # すべてのExcelファイルを別々のデータフレームにロード
# all_dfs = [pd.read_excel(file) for file in excel_files]

# # # すべてのデータフレームを一つに結合
# combined_df = pd.concat(all_dfs, ignore_index=True)
# print(combined_df)


# chromeブラウザ起動オプションを指定
chrome_options = webdriver.ChromeOptions()
# chrome_options.add_argument('--headless')
# chrome_options.add_argument('--no-sandbox')
# chrome_options.add_argument('--disable-dev-shm-usage')

# Chromeを立ち上げる
chrome_driver = webdriver.Chrome(options=chrome_options)


# WEB申請システムのトップページにアクセス
chrome_driver.get('https://xs016594.xsrv.jp/public/login')


# メールアドレスとパスワードを変数宣言する※ここをどうするか？テキストファイル（プロパティファイル）を別に設ける
email_address = 'test@example.com'
password = 'test01'

# メールアドレスとパスワードの入力欄を見つける
# parent_element = chrome_driver.find_element(
#     #    CSSセレクタを記述
#     By.CSS_SELECTOR, '[name="email"]')
email_input = chrome_driver.find_element(By.NAME, 'email')
password_input = chrome_driver.find_element(By.NAME, 'password')


# メールアドレスとパスワードを設定
email_input.send_keys(email_address)
password_input.send_keys(password)


# このボタンが表示されるまで30秒猶予をもたせる（※30秒後に操作するではない）
wait = WebDriverWait(chrome_driver, 30)

# ログインボタンをクリックしてログイン
form_login_button = wait.until(
    EC.visibility_of_element_located(
        (By.CSS_SELECTOR,
         #  ここを変更する
         '.mt-3.btn.webapplication-submit-button.w-100'
         )
    )
)

form_login_button.click()

# エクセルを立ち上げる
filepath = path = "/Users/kouheitakahashi/excel_auto/在庫管理表/在庫リスト.xlsx"

# Excelファイルを開く
wb = load_workbook(filename=filepath, data_only=True)

# 最初のシートを選択
sheet_obj = wb.active

# B列の各セルをチェック
print(sheet_obj.max_row)
shori = []
for row in range(4, sheet_obj.max_row + 1):
    print("welcome2")
    cell = sheet_obj['B{}'.format(row)]
    print(cell.value)
    # フラグが立っている場合は申請を行う
    if cell.value == 1 and sheet_obj['B13'].value != "申請済":
        # 実際に処理待つ
        time.sleep(5)

        # ログイン後、要素が読み込まれるまで待つ
        # wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'nav > img ')))

        # 申請メニュー画面上の操作
        # 備品・消耗品購入申請画面のリンクを見つける
        link = chrome_driver.find_element(
            By.XPATH, '//*[@id="app"]/main/div[1]/div/div[1]/div/a/h5')

        # リンクをクリックする
        link.click()

        time.sleep(15)
        # 1/18レッスンはここまで
        # 新規申請まで表示できた

        # 1/27のレッスンで実現したいこと
        # 承認画面のボタン操作までの流れを実装
        # # openpyexlを使っていく
        # １.指定フォルダ内にある在庫管理表にアクセス

        # Excelファイルへのパスを指定
        # path = "/Users/kouheitakahashi/excel_auto/在庫管理表/在庫リスト.xlsx"
        # # ワークブックとアクティブなシートを開く
        # wb_obj = openpyxl.load_workbook(path)
        # sheet_obj = wb_obj.active

        # 2.再発注用のフラグが立っているもの（B列）を参照する
        # （本来であればフラグを立っているものは全て申請するが今回は4行目のみ）
        # D列の4行目のセルを取得(商品名)
        cell_obj_item = sheet_obj.cell(row=row, column=4)
        # セルの値を出力
        print(cell_obj_item.value)

        # F列の4行目からのセルを取得(単価)
        cell_obj_price = sheet_obj.cell(row=row, column=6)
        # セルの値を出力
        print(cell_obj_price.value)

        # I列の4行目からのセルを取得(数量)
        cell_obj_quit = sheet_obj.cell(row=row, column=9)
        # セルの値を出力
        print(cell_obj_quit.value)

        # M列の4行目からのセルを取得(購入区分)
        cell_obj_purchase = sheet_obj.cell(row=row, column=13)
        # セルの値を出力
        print(cell_obj_purchase.value)

        # N列の4行目からのセルを取得(勘定科目)
        cell_obj_account = sheet_obj.cell(row=row, column=14)
        # セルの値を出力
        print(cell_obj_account.value)

        # 3.アクセス後４行目をWEB申請画面の新規申請に入力できるように設計する
        # （エクセルを参照して入力する動きを見たい）
        # エクセルから取得した2のデータをWEB申請システムの入力欄に入力
        # 指定されたXPathに一致する入力欄を見つけてデータを入力

        # 購入先
        input_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="app"]/main/div[1]/form/div[5]/input')
        input_field.send_keys('前回同様')  # ここに入力したいデータを指定

        # 購入先URL
        input_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="app"]/main/div[1]/form/div[6]/input')
        input_field.send_keys('前回同様')  # ここに入力したいデータを指定

        # 利用目的
        input_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="app"]/main/div[1]/form/div[7]/textarea')
        input_field.send_keys('前回同様')  # ここに入力したいデータを指定

        # 納品希望日
        current_date = datetime.now()
        one_week_later = current_date + timedelta(weeks=1)

        input_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="app"]/main/div[1]/form/div[8]/input')
        input_field.send_keys(
            one_week_later.strftime('00%Y%m%d'))  # ここに入力したいデータを指定

        # セレクトの場合の基本的な記述の仕方（区分）
        select_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="item_table"]/tbody/tr/td[1]/select')
        select = Select(select_field)
        print(select.options)
        select.select_by_value(str(cell_obj_purchase.value))

        # 商品名
        input_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="item_table"]/tbody/tr/td[2]/input')
        input_field.send_keys(cell_obj_item.value)  # ここに入力したいデータを指定

        # 購入単価
        input_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="item_table"]/tbody/tr/td[3]/input')
        input_field.send_keys(cell_obj_price.value)  # ここに入力したいデータを指定

        # 数量
        input_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="item_table"]/tbody/tr/td[4]/input')
        input_field.send_keys(cell_obj_quit.value)

        # 単位
        input_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="item_table"]/tbody/tr/td[5]/input')
        input_field.send_keys("式")

        # セレクトの場合の基本的な記述の仕方（勘定科目）
        select_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="item_table"]/tbody/tr/td[6]/select')
        select = Select(select_field)
        print(select.options)
        select.select_by_value(str(cell_obj_account.value))

        # スクロールを実行（例: 500ピクセル下にスクロール）
        chrome_driver.execute_script("window.scrollBy(0, 1500);")

        time.sleep(15)
        # メール送信操作（ボタン）
        # ボタンを選択
        button = chrome_driver.find_element(
            By.XPATH, '//*[@id="app"]/main/div[1]/form/button[1]')

        # ボタンをクリック
        button.click()

        time.sleep(3)

        # ベージボタンを選択
        button = chrome_driver.find_element(
            By.XPATH, '//*[@id="page-nate"]/button[3]')

        # ボタンをクリック
        button.click()

        time.sleep(3)

        # 送信先ボタンを選択
        button = chrome_driver.find_element(
            By.XPATH, '//*[@id="destination_body"]/tr[2]/td[5]/button')

        # ボタンをクリック
        button.click()

        time.sleep(3)

        # # メール送信（直接書き込む）
        # input_field = chrome_driver.find_element(
        #     By.XPATH,
        #     '//*[@id="email"]')
        # input_field.send_keys("cobo94251@gmail.com")

        # 備考
        input_field = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="app"]/main/div[1]/form/div[15]/textarea')
        input_field.send_keys("定量発注")

        time.sleep(3)
        # スクロールを実行（例: 500ピクセル下にスクロール）
        chrome_driver.execute_script("window.scrollBy(0, 1500);")

        time.sleep(5)
        # 4.承認ボタンを押し、承認する動作まで進める（申請できないが今回はそれで）
        # 指定されたXPathに一致するボタンを見つけてクリック
        button1 = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="app"]/main/div[1]/form/button[2]')
        button1.click()

        time.sleep(5)

        button2 = chrome_driver.find_element(
            By.XPATH,
            '//*[@id="exampleModal"]/div/div/div[3]/button[1]')
        button2.click()

        time.sleep(10)

        # # ブックを取得
        # ブック変数 = openpyxl.Workbook(ファイル名)
        # # シートを取得
        # 最初のシートを選択
        # sheet_obj = wb.active
        # セルへ書き込む
        # シート変数[セル記号] = 書き込む値
        # sheet_obj.cell(row=row, column=15).value = "申請済"
        # print(sheet_obj)
        shori.append(row)

        # wb.save("/Users/kouheitakahashi/excel_auto/在庫管理表/在庫リスト.xlsx")
wb.close()
# Excelファイルを開く
wb = load_workbook(filename=filepath)

# 最初のシートを選択
sheet_obj = wb.active

for row2 in shori:
    # シート変数[セル記号] = 書き込む値
    sheet_obj.cell(row=row2, column=15).value = "申請済"
    print(sheet_obj)

wb.save("/Users/kouheitakahashi/excel_auto/在庫管理表/在庫リスト.xlsx")
