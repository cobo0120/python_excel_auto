# 必要なライブラリをインポート
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from getpass import getpass
import time
import os
import pandas as pd
import glob
# 別のプログラム上でタイマー設定で起動（exeファイルを設定し起動）朝の9時に読み進める
# エクセルを見にいく（python）どこのエクセルを見にいくのか難しいため、あらかじめディレクトリ指定するか。exeファイルを作成するか。

# フォルダに保存されている指定のエクセルファイルを参照する


# def list_files_recursively(directory):
#     return glob.glob(os.path.join(directory, '**/*.xlsx'), recursive=True)


# path = '/Users/kouheitakahashi/excel_auto/在庫管理表/在庫リスト.xlsx'
# excel_files = list_files_recursively(path)

# # すべてのExcelファイルを別々のデータフレームにロード
# all_dfs = [pd.read_excel(file) for file in excel_files]

# # # すべてのデータフレームを一つに結合
# combined_df = pd.concat(all_dfs, ignore_index=True)
# print(combined_df)

# chromeブラウザ起動オプションを指定
chrome_options = webdriver.ChromeOptions()
# chrome_options.add_argument('--headless')
# chrome_options.add_argument('--no-sandbox')
# chrome_options.add_argument('--disable-dev-shm-usage')

# Chromeを立ち上げる
chrome_driver = webdriver.Chrome(options=chrome_options)


# WEB申請システムのトップページにアクセス
chrome_driver.get('https://xs016594.xsrv.jp/public/login')


# メールアドレスとパスワードを変数宣言する※ここをどうするか？テキストファイル（プロパティファイル）を別に設ける
email_address = 'test@example.com'
password = 'test01'

# メールアドレスとパスワードの入力欄を見つける
# parent_element = chrome_driver.find_element(
#     #    CSSセレクタを記述
#     By.CSS_SELECTOR, '[name="email"]')
email_input = chrome_driver.find_element(By.NAME, 'email')
password_input = chrome_driver.find_element(By.NAME, 'password')


# メールアドレスとパスワードを設定
email_input.send_keys(email_address)
password_input.send_keys(password)


# このボタンが表示されるまで30秒猶予をもたせる（※30秒後に操作するではない）
wait = WebDriverWait(chrome_driver, 30)

# ログインボタンをクリックしてログイン
form_login_button = wait.until(
    EC.visibility_of_element_located(
        (By.CSS_SELECTOR,
         #  ここを変更する
         '.mt-3.btn.webapplication-submit-button.w-100'
         )
    )
)

form_login_button.click()


# 実際に処理待つ
time.sleep(5)


# ログイン後、要素が読み込まれるまで待つ
# wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'nav > img ')))


# 申請メニュー画面上の操作
# 備品・消耗品購入申請画面のリンクを見つける
link = chrome_driver.find_element(
    By.XPATH, '//*[@id="app"]/main/div[1]/div/div[1]/div/a/h5')


# リンクをクリックする
link.click()

time.sleep(10)
# 1/18レッスンはここまで
# 新規申請まで表示できた


# 1/27のレッスンで実現したいこと
# 承認画面のボタン操作までの流れを実装
# # openpyexlを使っていく
# １.指定フォルダ内にある在庫管理表にアクセス

# Excelファイルへのパスを指定
path = "/Users/kouheitakahashi/excel_auto/在庫管理表/在庫リスト.xlsx"
# ワークブックとアクティブなシートを開く
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active


# 2.再発注用のフラグが立っているもの（B列）を参照する
# （本来であればフラグを立っているものは全て申請するが今回は4行目のみ）
# B列の4行目のセルを取得
cell_obj = sheet_obj.cell(row=4, column=2)
# セルの値を出力
print(cell_obj.value)

# 3.アクセス後４行目をWEB申請画面の新規申請に入力できるように設計する
# （エクセルを参照して入力する動きを見たいので少しの設定で１行入力で）
# 新しいデータを別の入力欄に入力
# 例えば、E列の4行目に入力する場合
sheet_obj.cell(row=4, column=5).value = cell_obj.value  # E列は5番目の列なのでcolumn=5


# 4.承認ボタンを押し、承認する動作まで進める（申請できないが今回はそれで）
# 指定されたXPathに一致するボタンを見つけてクリック
button1 = driver.find_element_by_xpath(
    '//*[@id="app"]/main/div[1]/form/button[2]')
button1.click()

button2 = driver.find_element_by_xpath(
    '//*[@id="exampleModal"]/div/div/div[3]/button[1]')
button2.click()
